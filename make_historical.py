## Added year2datetime function
## Made nutrient and toxin levels log-scaled
## interpolated throughout the year for each
## (no data between end of one year and start of next)


## Code to preprocess Detroit Lake empirical data Spring 2019
#! James Watson, The Prediction Lab 2019
import pandas as pd
import numpy as np
import openpyxl as px
import matplotlib.pylab as plt
from scipy import interpolate
interp_kind = 'zero' # 'previous'

###############################################################################
#### Function for calculating decimal years
from datetime import datetime
def yearLength(yr):
    yr = np.reshape(yr,(-1,1))
    year_length = [datetime(year=yr[k]+1, month=1, day=1) - datetime(year=yr[k], month=1, day=1) for k in np.arange(len(yr))]
    year_length = [year_length[k].days for k in np.arange(len(yr))]
    return year_length

def datetime2year(dt):
    year_part = dt - datetime(year=dt.year, month=1, day=1)
    year_length = datetime(year=dt.year+1, month=1, day=1) - datetime(year=dt.year, month=1, day=1)
    return dt.year + year_part/year_length

def year2datetime(date):
    year = np.array(np.uint(np.floor(date)), ndmin=2)
    year_part = date - year
    Seconds = year_part*np.uint(yearLength(year))*24*60*60
    Date = [ dt.datetime(year=year[0][k] ,month=1,day=1) + dt.timedelta(seconds = Seconds[0][k]) for k in np.arange(year.size) ]
    return Date

## Create daily timeseries in decimal years
import datetime as dt
date = dt.datetime(2013,1,1,12,0,0)
timeInt = dt.timedelta(days = 7)
TIME = datetime2year(date) # this is the daily timeseries
for i in range(320): # range(2200):
    date += timeInt
    TIME = np.hstack((TIME,datetime2year(date)))

## Define training and testing dates
## Locations we care about
LOCS = ['BB','BO','HA','HT','LB','LBP','LBS','MG']

## Make an interpolating function
def interpolator(times, data, mode, domain):
    f = interpolate.interp1d(times,data,kind=mode,bounds_error=False,fill_value=np.nan)
    return f(domain)

def integrator(times, data, mode, domain):
    dataInt = np.zeros(len(domain))
    for k in np.arange(1,len(domain)):
        mask = (times >= domain[k-1])&(times < domain[k])
        data = np.array(data)
        if np.any(mask):
            print(k)
            if mode == 'mean':
                dataMean = np.nanmean(data[mask])
                dataInt[k] = dataMean
            elif mode == 'max':
                dataMax = np.nanmax(data[mask])
                dataInt[k] = dataMax
            elif mode == 'min':
                dataMin = np.nanmin(data[mask])
                dataInt[k] = dataMin
            else:
                raise Exception('Not an appropriate input for integrator.')
                dataInt[k] = np.nan
        else:
            dataInt[k] = np.nan
    return dataInt

def seasonalCleaner(times, data, funct, mode, domain):
    t = np.array(times)
    N = funct(times, data, mode, domain)
    yr = 2013
    while yr <2019:
        print(str(yr)+': '+str(np.where((t>=yr)&(t<yr+1))[0].shape[0]) + ' records.')
        if (np.where(t < yr+1)[0].shape[0] > 0)&(np.where(t >= yr+1)[0].shape[0] > 0):
            maxSample = np.max(t[t < yr+1])
            nextSample = np.min(t[t >= yr+1])
            yl = yearLength(yr)
            yl = yl[0]
            maxInt = domain[np.min(np.where(domain >= maxSample)[0])]
            nextYr = int(np.floor(nextSample))
            yl = yearLength(nextYr)
            yl = yl[0]
            nextInt = domain[np.min(np.where(domain >= nextSample)[0])]
            N[(domain > maxInt)&(domain < nextInt)] = np.nan
            yr = nextYr
        else:
            yr += 1
    return N

################################################################ NUTRIENTS ####
data = px.load_workbook("../Data/Raw_lake/Historical/Nutrients.xlsx",data_only=True)
ws = data['Nutrients']

## Extract values
# time
column = ws['A'] 
Time = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    date = column[x].value
    Time[x-1] = datetime2year(date)

# site locations
column = ws['B'] 
loc = np.empty(len(column)-1,dtype='object')
for x in np.arange(1,len(column)):
    loc[x-1] = column[x].value

uloc = np.unique(loc)

# NO3+NO2 (mg/L)
column = ws['D'] 
nut1 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    nut1[x-1] = column[x].value #+ 0.0055
    try:
        nut1[x-1] = np.log( nut1[x-1] + 1 )
    except:
        nut1[x-1] = np.nan

# O-Phos (mg/L)
column = ws['E'] 
nut2 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    nut2[x-1] = column[x].value
    try:
        nut2[x-1] = np.log( nut2[x-1] + 1 )
    except:
        nut2[x-1] = np.nan

# TN (mg/L)
column = ws['F'] 
nut3 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    nut3[x-1] = column[x].value#+ 0.07 
    try:
        nut3[x-1] = np.log( nut3[x-1] + 1 )
    except:
        nut3[x-1] = np.nan

# T-Phos (mg/L)
column = ws['G'] 
nut4 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    nut4[x-1] = column[x].value# + 0.003
    try:
        nut4[x-1] = np.log( nut4[x-1] + 1 )
    except:
        nut4[x-1] = np.nan

# Interpolate to daily time series
# and rearrange so its timeseries of each nut for each location
NUT1 = np.ones((len(TIME),len(LOCS))) * np.nan
for i in np.arange(0,len(LOCS)): #1): #0,len(LOCS)):
    ID = np.where(loc == LOCS[i])[0]
    #ID = np.isin(loc, LOCS)
    if len(ID) > 1:
        n = nut1[ID]
        t = Time[ID] 
        print('Nutrient 1 (nitrate/nitrite):')
        N = seasonalCleaner(t,n,interpolator,interp_kind,TIME)
        NUT1[:,i] = N 
    
NUT2 = np.ones((len(TIME),len(LOCS))) * np.nan
for i in np.arange(1): #0,len(LOCS)):
    #ID = np.where(loc == LOCS[i])[0]
    ID = np.isin(loc, LOCS)
    if len(ID) > 3:
        n = nut2[ID]
        t = Time[ID] 
        print('Nutrient 2 (O-Phos):')
        N = seasonalCleaner(t,n,interpolator,interp_kind,TIME)
        NUT2[:,i] = N 
    
NUT3 = np.ones((len(TIME),len(LOCS))) * np.nan
for i in np.arange(1): #0,len(LOCS)):
    #ID = np.where(loc == LOCS[i])[0]
    ID = np.isin(loc, LOCS)
    if len(ID) > 1:
        n = nut3[ID]
        t = Time[ID] 
        print('Nutrient 3 (TN):')
        N = seasonalCleaner(t,n,interpolator,interp_kind,TIME)
        NUT3[:,i] = N 
    
NUT4 = np.ones((len(TIME),len(LOCS))) * np.nan
for i in np.arange(1): #0,len(LOCS)):
    #ID = np.where(loc == LOCS[i])[0]
    ID = np.isin(loc, LOCS)
    if len(ID) > 1:
        n = nut4[ID]
        t = Time[ID] 
        print('Nutrient 4 (T-Phos):')
        N = seasonalCleaner(t,n,interpolator,interp_kind,TIME)
        NUT4[:,i] = N 
    


################################################################### TOXINS ####
 
data = px.load_workbook("../Data/Raw_lake/Historical/CyanotoxinConcentrations.xlsx",data_only=True)
## Extract values from LCMSMS
# Time
ws = data['LCMSMS']
column = ws['A']
time = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    date = column[x].value
    time[x-1] = datetime2year(date)

# Locations
ws = data['LCMSMS']
column = ws['B']
loc = np.empty(len(column)-1,dtype='object')
for x in np.arange(1,len(column)):
    loc[x-1] = np.str(column[x].value)

uloc = np.unique(loc)

# Cylindro (ppb)
column = ws['C']
tox1 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    tox1[x-1] = column[x].value
    try:
        tox1[x-1] = np.log( tox1[x-1] )
    except:
        tox1[x-1] = np.nan

# Microcystin (ppb)
column = ws['D']
tox2 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    tox2[x-1] = column[x].value
    try:
        tox2[x-1] = np.log(tox2[x-1])
    except:
        tox2[x-1] = np.nan

# Interpolate to daily time series
# and rearrange so its timeseries of each nut for each location
TOX1 = np.ones((len(TIME),len(LOCS))) * np.nan
for i in np.arange(1): #0,len(LOCS)):
    #ID = np.where(loc == LOCS[i])[0]
    ID = np.isin(loc, LOCS)
    if len(ID)>3:
        n = tox1[ID]
        t = time[ID]
        print('LC Cylindro:')
        N = seasonalCleaner(t,n,interpolator,interp_kind,TIME)
        TOX1[:,i] = N

TOX2 = np.ones((len(TIME),len(LOCS))) * np.nan
for i in np.arange(1): #0,len(LOCS)):
    #ID = np.where(loc == LOCS[i])[0]
    ID = np.isin(loc, LOCS)
    if len(ID)>3:
        n = tox2[ID]
        t = time[ID]
        print('LC Microcystin:')
        N = seasonalCleaner(t,n,interpolator,interp_kind,TIME)
        TOX2[:,i] = N


## Extract values from ELISA
# Time
ws = data['ELISA']
column = ws['A']
time = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    date = column[x].value
    time[x-1] = datetime2year(date)

# Locations
ws = data['ELISA']
column = ws['B']
loc = np.empty(len(column)-1,dtype='object')
for x in np.arange(1,len(column)):
    loc[x-1] = np.str(column[x].value)

uloc = np.unique(loc)

# Cylindro (ppb)
column = ws['C']
tox3 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    tox3[x-1] = column[x].value
    try:
        tox3[x-1] = np.log(tox3[x-1])
    except:
        tox3[x-1] = np.nan

# Microcystin (ppb)
column = ws['D']
tox4 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    tox4[x-1] = column[x].value
    try:
        tox4[x-1] = np.log(tox4[x-1])
    except:
        tox4[x-1] = np.nan

# Interpolate to daily time series
# and rearrange so its timeseries of each nut for each location
TOX3 = np.ones((len(TIME),len(LOCS))) * np.nan
for i in np.arange(0,len(LOCS)):#1): #0,len(LOCS)):
    ID = np.where(loc == LOCS[i])[0]
    #ID = np.isin(loc, LOCS)
    if len(ID)>1:
        n = tox3[ID]
        t = time[ID]
        print('ELISA Cylindro:')
        N = seasonalCleaner(t,n,interpolator,interp_kind,TIME)
        TOX3[:,i] = N

TOX4 = np.ones((len(TIME),len(LOCS))) * np.nan
for i in np.arange(1): #0,len(LOCS)):
    #ID = np.where(loc == LOCS[i])[0]
    ID = np.isin(loc, LOCS)
    if len(ID)>3:
        n = tox4[ID]
        t = time[ID]
        print('ELISA Microcystin:')
        N = seasonalCleaner(t,n,interpolator,interp_kind,TIME)
        TOX4[:,i] = N


################################################################ ALGAE  ####
data = px.load_workbook("../Data/Raw_lake/Historical/Algae Speciation.xlsx",data_only=True)
ws = data['PrioritySites']

## Extract values
# time
columnd = ws['B'] 
columnt = ws['C']
time = np.zeros(len(columnd)-1)
for x in np.arange(1,len(columnd)):
    dated = columnd[x].value
    datet = columnt[x].value
    try:
        date = dt.datetime.combine(dated,datet)
    except:
        date = dated
    time[x-1] = datetime2year(date)

# site locations
column = ws['A'] 
loc = np.empty(len(column)-1,dtype='object')
for x in np.arange(1,len(column)):
    loc[x-1] = column[x].value

uloc = np.unique(loc)

# Genus
column = ws['E'] 
gen = np.empty(len(column)-1,dtype='object')
for x in np.arange(1,len(column)):
    gen[x-1] = np.str(column[x].value)

# Division
column = ws['F'] 
div = np.empty(len(column)-1,dtype='object')
for x in np.arange(1,len(column)):
    div[x-1] = np.str(column[x].value)
udiv = np.unique(div)
DIV = udiv

# Tally
column = ws['G'] 
tal = np.zeros(len(column)-1) * np.nan
for x in np.arange(1,len(column)):
    tal[x-1] = column[x].value

# Density
column = ws['I'] 
den = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    den[x-1] = column[x].value

# Tot Biovolume
column = ws['K'] 
tbv = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    tbv[x-1] = column[x].value
    try:
        tbv[x-1] = np.log(tbv[x-1])
    except:
        tbv[x-1] = np.nan
    #tbv[x-1] = column[x].value

# % Biovolume
column = ws['L'] 
fbv = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    fbv[x-1] = column[x].value

# Interpolate time series
# and rearrange so its timeseries of each div
DEN = np.ones((len(TIME),len(udiv))) * np.nan # DEN = np.ones((len(TIME),len(LOCS),len(udiv))) * np.nan
for i in np.arange(1): #0,len(LOCS)):
    #ID = np.where(loc == LOCS[i])[0]
    ID = np.isin(loc, LOCS)
    if len(ID)>1:
        for j in np.arange(0,len(udiv)):
            KD = np.where((div == udiv[j]) & ID)[0] # (loc == LOCS[i]))[0]
            if len(KD) > 1:
                n = den[KD]
                t = time[KD]
                n1 = []
                t1 = []
                tvals = np.unique(t)
                for s in np.arange(len(tvals)):
                    conc_div_samples = np.where(t == tvals[s])[0]
                    n1.append(np.nansum(n[conc_div_samples]))
                    t1.append(tvals[s])
                if (len(n1) != len(tvals))|(len(t1) != len(tvals)):
                    raise Exception("Cell density data compilation failed.")
                N = seasonalCleaner(t1,n1,integrator,'mean',TIME)
                DEN[:,j] = N

# TBV = np.ones((len(TIME),len(LOCS),len(udiv))) * np.nan
## make sure to sum across concurrent samples (same div, different genus)
## then average across locations.
TBV = np.ones((len(TIME),len(udiv))) * np.nan
for i in np.arange(1): #0,len(LOCS)):
    #ID = np.where(loc == LOCS[i])[0]
    ID = np.isin(loc, LOCS)
    if len(ID)>1:
        for j in np.arange(0,len(udiv)):
            KD = np.where((div == udiv[j]) & ID)[0]#(loc == LOCS[i]))[0]
            if len(KD) > 1:
                n = tbv[KD]
                t = time[KD]
                n1 = []
                t1 = []
                tvals = np.unique(t)
                for s in np.arange(len(tvals)):
                    conc_div_samples = np.where(t == tvals[s])[0]
                    n1.append(np.nansum(n[conc_div_samples]))
                    t1.append(tvals[s])
                if (len(n1) != len(tvals))|(len(t1) != len(tvals)):
                    raise Exception("Biovolume data compilation failed.")
                print('Total biovolume of div '+str(j))
                N = seasonalCleaner(t1,n1,integrator,'mean',TIME)
                #TBV[:,i,j] = N
                TBV[:,j] = N

FBV = np.ones((len(TIME),len(LOCS),len(udiv))) * np.nan
for i in np.arange(1): #0,len(LOCS)):
    #ID = np.where(loc == LOCS[i])[0]
    ID = np.isin(loc, LOCS)
    if len(ID)>1:
        for j in np.arange(0,len(udiv)):
            KD = np.where((div == udiv[j]) & ID)[0] # (loc == LOCS[i]))[0]
            if len(KD) > 1:
                n = fbv[KD]
                t = time[KD]
                n1 = []
                t1 = []
                tvals = np.unique(t)
                for s in np.arange(len(tvals)):
                    conc_div_samples = np.where(t == tvals[s])[0]
                    n1.append(np.nansum(n[conc_div_samples]))
                    t1.append(tvals[s])
                if (len(n1) != len(tvals))|(len(t1) != len(tvals)):
                    raise Exception("Fractional biovolume data compilation failed.")
                print('Frac biovolume of div '+str(j))
                N = seasonalCleaner(t1,n1,interpolator,interp_kind,TIME)
                f = interpolate.interp1d(t,n,kind=interp_kind,bounds_error=False,fill_value=np.nan)
                FBV[:,i,j] = N


################################################################ WEATHER ####
data = px.load_workbook("../Data/Raw_lake/Historical/Weather data.xlsx",data_only=True)
ws = data['Weather-BureauRecl Detroit Lake']

## Extract values
# time
column = ws['A'] 
time = np.zeros(len(column)-2)
for x in np.arange(1,len(column)-1):
    date = column[x].value
    time[x-1] = datetime2year(date)

# temperature
column = ws['B'] 
tem = np.ones(len(column)-2) 
for x in np.arange(1,len(column)-1):
    y = column[x].value
    if type(y) == float:
        tem[x-1] = y

# humidity
column = ws['C'] 
hum = np.ones(len(column)-2) 
for x in np.arange(1,len(column)-1):
    y = column[x].value
    if type(y) == float:
        hum[x-1] = y

# peak wind
column = ws['F'] 
pwi = np.ones(len(column)-2) 
for x in np.arange(1,len(column)-1):
    y = column[x].value
    if type(y) == float:
        pwi[x-1] = y

# wind speed
column = ws['G'] 
wis = np.ones(len(column)-2) 
for x in np.arange(1,len(column)-1):
    y = column[x].value
    if type(y) == float:
        wis[x-1] = y

# rain
column = ws['I'] 
rain = np.ones(len(column)-2) 
for x in np.arange(1,len(column)-1):
    y = column[x].value
    if type(y) == float:
        rain[x-1] = y

# pressure
column = ws['J'] 
pres = np.ones(len(column)-2) 
for x in np.arange(1,len(column)-1):
    y = column[x].value
    if type(y) == float:
        pres[x-1] = y

# Interpolate to daily time series
# and rearrange so its timeseries of each nut for each location
# f = interpolate.interp1d(time,tem,kind=interp_kind,bounds_error=False,fill_value=np.nan)
# N = f(TIME)
# JD = np.where(np.isnan(N)==0)[0]
# KD = np.where(np.diff(N[JD])==0)
# N[JD[KD]] = np.nan
# TEMP = N / 10

N = seasonalCleaner(time, tem, integrator, 'mean', TIME)
TEMP = N / 10

#N = seasonalCleaner(time, np.power(tem/10,2), integrator, 'mean', TIME)
#T2 = N

N = seasonalCleaner(time, tem, integrator, 'min', TIME)
MINT = N / 10

N = seasonalCleaner(time, tem, integrator, 'max', TIME)
MAXT = N / 10

f = interpolate.interp1d(time,hum,kind=interp_kind,bounds_error=False,fill_value=np.nan)
N = f(TIME)
JD = np.where(np.isnan(N)==0)[0]
KD = np.where(np.diff(N[JD])==0)    
N[JD[KD]] = np.nan
HUM = N

f = interpolate.interp1d(time,pwi,kind=interp_kind,bounds_error=False,fill_value=np.nan)
N = f(TIME)
JD = np.where(np.isnan(N)==0)[0]
KD = np.where(np.diff(N[JD])==0)
N[JD[KD]] = np.nan
PWI = N

f = interpolate.interp1d(time,wis,kind=interp_kind,bounds_error=False,fill_value=np.nan)
N = f(TIME)
JD = np.where(np.isnan(N)==0)[0]
KD = np.where(np.diff(N[JD])==0)
N[JD[KD]] = np.nan
WIS = N

# Make a degree day count - # of daily max temps exceeding a given value
tempThreshold = 65

## Get daily time series, divided at midnight
date = dt.datetime(2013,1,1,0,0,0)
timeInt = dt.timedelta(days = 1)
TIMED = datetime2year(date) # this is the daily timeseries
for i in range(2200):
    date += timeInt
    TIMED = np.hstack((TIMED,datetime2year(date)))

## Get daily max temp over previous day
N = seasonalCleaner(time, tem, integrator, 'max', TIMED)
DEG = [np.nanmax([N[k]-tempThreshold,0]) for k in np.arange(len(N))]
DDEG = np.zeros(len(DEG))
for yr in np.arange(2013,2020):
    days = (TIMED >= yr) & (TIMED < yr+1)
    DDEG[days] = np.cumsum(np.array(DEG)[days])

## Get weekly cumulative degree days
N = seasonalCleaner(TIMED,DDEG,integrator, 'mean', TIME)
WDEG = N*7
## And rescale
WDEG = WDEG/10000

N = seasonalCleaner(time,rain,integrator,'max',TIME)
RAIN = N

N = seasonalCleaner(time,pres,integrator,'mean',TIME)
PRES = N

# Calculate species diversity
#DEN = np.nansum(DEN,1)
PRCT = DEN/np.repeat(np.reshape(np.sum(DEN,1),(-1,1)), DEN.shape[1], 1)
LPRCT = np.log(PRCT)
LPRCT[np.where(LPRCT==-np.inf)] = 0
DVR = np.sum(PRCT * LPRCT,1)

################################################################ SAVE ########
# Locations: LOCS = ['BB','BO','HA','HT','LB','LBP','LBS']
# TIME: decimal years (daily increments, at midday)
# NUT1: NO3+NO2 (mg/L)                                  sparse
# NUT2: O-Phos (mg/L)                                   sparse
# NUT3: TN (mg/L)                                       sparse
# NUT4: T-Phos (mg/L)                                   sparse
# TOX1: LCMSMS Cylindro (ppb)                           sparse
# TOX2: LCMSMS Microcystin (ppb)
# TOX3: ELISA Cylindro (ppb)                            sparse
# TOX4: ELISA Microcystin (ppb)
# DIV: bacterial/algal family name
# DEN: algal concentration
# TBV: total biovolume
# FBV: fractional biovolume
# TEMP: temperature 
# T2: temp squared
# MINT: minimum temperature over the period
# MAXT: maximum temperature over the period
# HUM: humidity
# PWI: peak wind speed
# WIS: wind speed
# RAIN: rain
# PRES: barometric pressure
# WDEG: degree days integrated over each period
# DVR: shannon diversity of algal populations           sparse
np.savez("../Data/Preprocessed/Data_hist_interp.npz",LOCS=LOCS,TIME=TIME,NUT1=NUT1,WDEG=WDEG,
        NUT2=NUT2,NUT3=NUT3,NUT4=NUT4,TOX1=TOX1,TOX2=TOX2,TOX3=TOX3,TOX4=TOX4,DIV=DIV,DEN=DEN,DVR=DVR,
        TBV=TBV,FBV=FBV,TEMP=TEMP,MINT=MINT,MAXT=MAXT,HUM=HUM,PWI=PWI,WIS=WIS,RAIN=RAIN,PRES=PRES)


