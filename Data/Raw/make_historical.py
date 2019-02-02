## Code to preprocess Detroit Lake empirical data Spring 2019
#! James Watson, The Prediction Lab 2019
import pandas as pd
import numpy as np
import openpyxl as px
import matplotlib.pylab as plt
from scipy import interpolate


###############################################################################
#### Function for calculating decimal years
from datetime import datetime
def datetime2year(dt):
    year_part = dt - datetime(year=dt.year, month=1, day=1)
    year_length = datetime(year=dt.year+1, month=1, day=1) - datetime(year=dt.year, month=1, day=1)
    return dt.year + year_part/year_length

## Create daily timeseries in decimal years
import datetime as dt
date = dt.datetime(2013,1,1,12,0,0)
TIME = datetime2year(date) # this is the daily timeseries
for i in range(2100):
    date += dt.timedelta(days=1)
    TIME = np.hstack((TIME,datetime2year(date)))

## Locations we care about
LOCS = ['BB','BO','HA','HT','LB','LBP','LBS']


################################################################ NUTRIENTS ####
data = px.load_workbook("./Historical/Nutrients.xlsx",data_only=True)
ws = data['Nutrients']

## Extract values
# time
column = ws['A'] 
Time = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    date = column[x].value
    Time[x-1] = datetime2year(date)

# site locations
column = ws['B'] 
loc = np.empty(len(column)-1,dtype='object')
for x in np.arange(1,len(column)):
    loc[x-1] = column[x].value
uloc = np.unique(loc)

# NO3+NO2 (mg/L)
column = ws['D'] 
nut1 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    nut1[x-1] = column[x].value

# O-Phos (mg/L)
column = ws['E'] 
nut2 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    nut2[x-1] = column[x].value

# TN (mg/L)
column = ws['F'] 
nut3 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    nut3[x-1] = column[x].value

# T-Phos (mg/L)
column = ws['G'] 
nut4 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    nut4[x-1] = column[x].value

# Interpolate to daily time series
# and rearrange so its timeseries of each nut for each location
NUT1 = np.ones((len(TIME),len(LOCS))) * np.nan
for i in np.arange(0,len(LOCS)):
    ID = np.where(loc == LOCS[i])[0]
    n = nut1[ID]
    t = Time[ID] 
    f = interpolate.interp1d(t,n,kind='nearest',bounds_error=False,fill_value=np.nan)
    N = f(TIME) 
    JD = np.where(np.isnan(N)==0)[0]
    KD = np.where(np.diff(N[JD])==0)
    N[JD[KD]] = np.nan
    NUT1[:,i] = N 
    
NUT2 = np.ones((len(TIME),len(LOCS))) * np.nan
for i in np.arange(0,len(LOCS)):
    ID = np.where(loc == LOCS[i])[0]
    n = nut2[ID]
    t = Time[ID] 
    f = interpolate.interp1d(t,n,kind='nearest',bounds_error=False,fill_value=np.nan)
    N = f(TIME) 
    JD = np.where(np.isnan(N)==0)[0]
    KD = np.where(np.diff(N[JD])==0)
    N[JD[KD]] = np.nan
    NUT2[:,i] = N 
    
NUT3 = np.ones((len(TIME),len(LOCS))) * np.nan
for i in np.arange(0,len(LOCS)):
    ID = np.where(loc == LOCS[i])[0]
    n = nut3[ID]
    t = Time[ID] 
    f = interpolate.interp1d(t,n,kind='nearest',bounds_error=False,fill_value=np.nan)
    N = f(TIME) 
    JD = np.where(np.isnan(N)==0)[0]
    KD = np.where(np.diff(N[JD])==0)
    N[JD[KD]] = np.nan
    NUT3[:,i] = N 
    
NUT4 = np.ones((len(TIME),len(LOCS))) * np.nan
for i in np.arange(0,len(LOCS)):
    ID = np.where(loc == LOCS[i])[0]
    n = nut4[ID]
    t = Time[ID] 
    f = interpolate.interp1d(t,n,kind='nearest',bounds_error=False,fill_value=np.nan)
    N = f(TIME) 
    JD = np.where(np.isnan(N)==0)[0]
    KD = np.where(np.diff(N[JD])==0)
    N[JD[KD]] = np.nan
    NUT4[:,i] = N 
    


################################################################### TOXINS ####
data = px.load_workbook("./Historical/CyanotoxinConcentrations.xlsx",data_only=True)

# Extract values from LCMSMS
ws = data['LCMSMS']
column = ws['A']
Time = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    date = column[x].value
    Time[x-1] = datetime2year(date)

column = ws['C'] # Cylindro (ppb)
tox1 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    tox1[x-1] = column[x].value

column = ws['D'] # Microcystin (ppb)
tox2 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    tox2[x-1] = column[x].value

# Interpolate to daily time series
from scipy import interpolate
f1 = interpolate.interp1d(Time, tox1,kind='linear',bounds_error=False,fill_value=np.nan)
f2 = interpolate.interp1d(Time, tox2,kind='linear',bounds_error=False,fill_value=np.nan)

# These are the daily nutrient timeseries to save
TOX1 = f1(TIME) 
TOX2 = f2(TIME)


# Extract values from ELISA
ws = data['ELISA']
column = ws['A']
Time = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    date = column[x].value
    Time[x-1] = datetime2year(date)

column = ws['C'] # Cylindro (ppb)
tox1 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    tox1[x-1] = column[x].value

column = ws['D'] # Microcystin (ppb)
tox2 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    tox2[x-1] = column[x].value

# Interpolate to daily time series
from scipy import interpolate
f1 = interpolate.interp1d(Time, tox1,kind='linear',bounds_error=False,fill_value=np.nan)
f2 = interpolate.interp1d(Time, tox2,kind='linear',bounds_error=False,fill_value=np.nan)

# These are the daily nutrient timeseries to save
TOX3 = f1(TIME) 
TOX4 = f2(TIME)





