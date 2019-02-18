## Code to preprocess Detroit Lake empirical data Spring 2019
#! James Watson, The Prediction Lab 2019
import pandas as pd
import numpy as np
import openpyxl as px
import matplotlib.pylab as plt
from scipy import interpolate


###############################################################################
#### Function for calculating decimal years
from datetime import datetime
def datetime2year(dt):
    year_part = dt - datetime(year=dt.year, month=1, day=1)
    year_length = datetime(year=dt.year+1, month=1, day=1) - datetime(year=dt.year, month=1, day=1)
    return dt.year + year_part/year_length

## Create daily timeseries in decimal years
import datetime as dt
date = dt.datetime(2013,1,1,12,0,0)
TIME = datetime2year(date) # this is the daily timeseries
for i in range(round(2200/7)):
    date += dt.timedelta(days=7) ##<<<<<<<<< Change the temporal resolution of the interpolation
    TIME = np.hstack((TIME,datetime2year(date)))

## Locations we care about
LOCS = ['BB','BO','HA','HT','LB','LBP','LBS']


################################################################ NUTRIENTS ####
data = px.load_workbook("./Historical/Nutrients.xlsx",data_only=True)
ws = data['Nutrients']

## Extract values
# time
column = ws['A'] 
Time = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    date = column[x].value
    Time[x-1] = datetime2year(date)

# site locations
column = ws['B'] 
loc = np.empty(len(column)-1,dtype='object')
for x in np.arange(1,len(column)):
    loc[x-1] = column[x].value
uloc = np.unique(loc)

# NO3+NO2 (mg/L)
column = ws['D'] 
nut1 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    nut1[x-1] = column[x].value

# O-Phos (mg/L)
column = ws['E'] 
nut2 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    nut2[x-1] = column[x].value

# TN (mg/L)
column = ws['F'] 
nut3 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    nut3[x-1] = column[x].value

# T-Phos (mg/L)
column = ws['G'] 
nut4 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    nut4[x-1] = column[x].value

# Interpolate to daily time series
# and rearrange so its timeseries of each nut for each location
NUT1 = np.ones((len(TIME),len(LOCS))) * np.nan
for i in np.arange(0,len(LOCS)):
    ID = np.where(loc == LOCS[i])[0]
    n = nut1[ID]
    t = Time[ID] 
    f = interpolate.interp1d(t,n,kind='nearest',bounds_error=False,fill_value=np.nan)
    N = f(TIME) 
    JD = np.where(np.isnan(N)==0)[0]
    KD = np.where(np.diff(N[JD])==0)
    N[JD[KD]] = np.nan
    NUT1[:,i] = N 
    
NUT2 = np.ones((len(TIME),len(LOCS))) * np.nan
for i in np.arange(0,len(LOCS)):
    ID = np.where(loc == LOCS[i])[0]
    n = nut2[ID]
    t = Time[ID] 
    f = interpolate.interp1d(t,n,kind='nearest',bounds_error=False,fill_value=np.nan)
    N = f(TIME) 
    JD = np.where(np.isnan(N)==0)[0]
    KD = np.where(np.diff(N[JD])==0)
    N[JD[KD]] = np.nan
    NUT2[:,i] = N 
    
NUT3 = np.ones((len(TIME),len(LOCS))) * np.nan
for i in np.arange(0,len(LOCS)):
    ID = np.where(loc == LOCS[i])[0]
    n = nut3[ID]
    t = Time[ID] 
    f = interpolate.interp1d(t,n,kind='nearest',bounds_error=False,fill_value=np.nan)
    N = f(TIME) 
    JD = np.where(np.isnan(N)==0)[0]
    KD = np.where(np.diff(N[JD])==0)
    N[JD[KD]] = np.nan
    NUT3[:,i] = N 
    
NUT4 = np.ones((len(TIME),len(LOCS))) * np.nan
for i in np.arange(0,len(LOCS)):
    ID = np.where(loc == LOCS[i])[0]
    n = nut4[ID]
    t = Time[ID] 
    f = interpolate.interp1d(t,n,kind='nearest',bounds_error=False,fill_value=np.nan)
    N = f(TIME) 
    JD = np.where(np.isnan(N)==0)[0]
    KD = np.where(np.diff(N[JD])==0)
    N[JD[KD]] = np.nan
    NUT4[:,i] = N 
    


################################################################### TOXINS ####
data = px.load_workbook("./Historical/CyanotoxinConcentrations.xlsx",data_only=True)

## Extract values from LCMSMS
# Time
ws = data['LCMSMS']
column = ws['A']
time = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    date = column[x].value
    time[x-1] = datetime2year(date)

# Locations
ws = data['LCMSMS']
column = ws['B']
loc = np.empty(len(column)-1,dtype='object')
for x in np.arange(1,len(column)):
    loc[x-1] = np.str(column[x].value)
uloc = np.unique(loc)

# Cylindro (ppb)
column = ws['C']
tox1 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    tox1[x-1] = column[x].value

# Microcystin (ppb)
column = ws['D']
tox2 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    tox2[x-1] = column[x].value

# Interpolate to daily time series
# and rearrange so its timeseries of each nut for each location
TOX1 = np.ones((len(TIME),len(LOCS))) * np.nan
for i in np.arange(0,len(LOCS)):
    ID = np.where(loc == LOCS[i])[0]
    if len(ID)>3:
        n = tox1[ID]
        t = time[ID]
        f = interpolate.interp1d(t,n,kind='nearest',bounds_error=False,fill_value=np.nan)
        N = f(TIME)
        JD = np.where(np.isnan(N)==0)[0]
        KD = np.where(np.diff(N[JD])==0)
        N[JD[KD]] = np.nan
        TOX1[:,i] = N

TOX2 = np.ones((len(TIME),len(LOCS))) * np.nan
for i in np.arange(0,len(LOCS)):
    ID = np.where(loc == LOCS[i])[0]
    if len(ID)>3:
        n = tox2[ID]
        t = time[ID]
        f = interpolate.interp1d(t,n,kind='nearest',bounds_error=False,fill_value=np.nan)
        N = f(TIME)
        JD = np.where(np.isnan(N)==0)[0]
        KD = np.where(np.diff(N[JD])==0)
        N[JD[KD]] = np.nan
        TOX2[:,i] = N


## Extract values from ELISA
# Time
ws = data['ELISA']
column = ws['A']
time = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    date = column[x].value
    time[x-1] = datetime2year(date)

# Locations
ws = data['ELISA']
column = ws['B']
loc = np.empty(len(column)-1,dtype='object')
for x in np.arange(1,len(column)):
    loc[x-1] = np.str(column[x].value)
uloc = np.unique(loc)

# Cylindro (ppb)
column = ws['C']
tox3 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    tox3[x-1] = column[x].value

# Microcystin (ppb)
column = ws['D']
tox4 = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    tox4[x-1] = column[x].value

# Interpolate to daily time series
# and rearrange so its timeseries of each nut for each location
TOX3 = np.ones((len(TIME),len(LOCS))) * np.nan
for i in np.arange(0,len(LOCS)):
    ID = np.where(loc == LOCS[i])[0]
    if len(ID)>3:
        n = tox3[ID]
        t = time[ID]
        f = interpolate.interp1d(t,n,kind='nearest',bounds_error=False,fill_value=np.nan)
        N = f(TIME)
        JD = np.where(np.isnan(N)==0)[0]
        KD = np.where(np.diff(N[JD])==0)
        N[JD[KD]] = np.nan
        TOX3[:,i] = N

TOX4 = np.ones((len(TIME),len(LOCS))) * np.nan
for i in np.arange(0,len(LOCS)):
    ID = np.where(loc == LOCS[i])[0]
    if len(ID)>3:
        n = tox4[ID]
        t = time[ID]
        f = interpolate.interp1d(t,n,kind='nearest',bounds_error=False,fill_value=np.nan)
        N = f(TIME)
        JD = np.where(np.isnan(N)==0)[0]
        KD = np.where(np.diff(N[JD])==0)
        N[JD[KD]] = np.nan
        TOX4[:,i] = N


################################################################ ALGAE  ####
data = px.load_workbook("./Historical/Algae Speciation.xlsx",data_only=True)
ws = data['PrioritySites']

## Extract values
# time
column = ws['B'] 
time = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    date = column[x].value
    time[x-1] = datetime2year(date)

# site locations
column = ws['A'] 
loc = np.empty(len(column)-1,dtype='object')
for x in np.arange(1,len(column)):
    loc[x-1] = column[x].value
uloc = np.unique(loc)

# Genus
column = ws['E'] 
gen = np.empty(len(column)-1,dtype='object')
for x in np.arange(1,len(column)):
    gen[x-1] = np.str(column[x].value)

# Division
column = ws['F'] 
div = np.empty(len(column)-1,dtype='object')
for x in np.arange(1,len(column)):
    div[x-1] = np.str(column[x].value)
udiv = np.unique(div)
DIV = udiv

# Tally
column = ws['G'] 
tal = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    tal[x-1] = column[x].value

# Density
column = ws['I'] 
den = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    den[x-1] = column[x].value

# Tot Biovolumn
column = ws['K'] 
tbv = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    tbv[x-1] = column[x].value

# % Biovolumn
column = ws['L'] 
fbv = np.zeros(len(column)-1)
for x in np.arange(1,len(column)):
    fbv[x-1] = column[x].value

# Interpolate to daily time series
# and rearrange so its timeseries of each nut for each location
DEN = np.ones((len(TIME),len(LOCS),len(udiv))) * np.nan
for i in np.arange(0,len(LOCS)):
    ID = np.where(loc == LOCS[i])[0]
    if len(ID)>3:
        for j in np.arange(0,len(udiv)):
            KD = np.where((div == udiv[j]) & (loc == LOCS[i]))[0]
            if len(KD) > 1:
                n = den[KD]
                t = time[KD]
                f = interpolate.interp1d(t,n,kind='nearest',bounds_error=False,fill_value=np.nan)
                N = f(TIME)
                JD = np.where(np.isnan(N)==0)[0]
                KD = np.where(np.diff(N[JD])==0)
                N[JD[KD]] = np.nan
                DEN[:,i,j] = N

TBV = np.ones((len(TIME),len(LOCS),len(udiv))) * np.nan
for i in np.arange(0,len(LOCS)):
    ID = np.where(loc == LOCS[i])[0]
    if len(ID)>3:
        for j in np.arange(0,len(udiv)):
            KD = np.where((div == udiv[j]) & (loc == LOCS[i]))[0]
            if len(KD) > 1:
                n = tbv[KD]
                t = time[KD]
                f = interpolate.interp1d(t,n,kind='nearest',bounds_error=False,fill_value=np.nan)
                N = f(TIME)
                JD = np.where(np.isnan(N)==0)[0]
                KD = np.where(np.diff(N[JD])==0)
                N[JD[KD]] = np.nan
                TBV[:,i,j] = N

FBV = np.ones((len(TIME),len(LOCS),len(udiv))) * np.nan
for i in np.arange(0,len(LOCS)):
    ID = np.where(loc == LOCS[i])[0]
    if len(ID)>3:
        for j in np.arange(0,len(udiv)):
            KD = np.where((div == udiv[j]) & (loc == LOCS[i]))[0]
            if len(KD) > 1:
                n = fbv[KD]
                t = time[KD]
                f = interpolate.interp1d(t,n,kind='nearest',bounds_error=False,fill_value=np.nan)
                N = f(TIME)
                JD = np.where(np.isnan(N)==0)[0]
                KD = np.where(np.diff(N[JD])==0)
                N[JD[KD]] = np.nan
                FBV[:,i,j] = N


################################################################ WEATHER ####
data = px.load_workbook("./Historical/Weather data.xlsx",data_only=True)
ws = data['Weather-BureauRecl Detroit Lake']

## Extract values
# time
column = ws['A'] 
time = np.zeros(len(column)-2)
for x in np.arange(1,len(column)-1):
    date = column[x].value
    time[x-1] = datetime2year(date)

# temperature
column = ws['B'] 
tem = np.ones(len(column)-2) 
for x in np.arange(1,len(column)-1):
    y = column[x].value
    if type(y) == float:
        tem[x-1] = y

# humidity
column = ws['C'] 
hum = np.ones(len(column)-2) 
for x in np.arange(1,len(column)-1):
    y = column[x].value
    if type(y) == float:
        hum[x-1] = y

# peak wind
column = ws['F'] 
pwi = np.ones(len(column)-2) 
for x in np.arange(1,len(column)-1):
    y = column[x].value
    if type(y) == float:
        pwi[x-1] = y

# wind speed
column = ws['G'] 
wis = np.ones(len(column)-2) 
for x in np.arange(1,len(column)-1):
    y = column[x].value
    if type(y) == float:
        wis[x-1] = y

# rain
column = ws['I'] 
rain = np.ones(len(column)-2) 
for x in np.arange(1,len(column)-1):
    y = column[x].value
    if type(y) == float:
        rain[x-1] = y

# pressure
column = ws['J'] 
pres = np.ones(len(column)-2) 
for x in np.arange(1,len(column)-1):
    y = column[x].value
    if type(y) == float:
        pres[x-1] = y

# Interpolate to daily time series
# and rearrange so its timeseries of each nut for each location
f = interpolate.interp1d(time,tem,kind='nearest',bounds_error=False,fill_value=np.nan)
N = f(TIME)
JD = np.where(np.isnan(N)==0)[0]
KD = np.where(np.diff(N[JD])==0)
N[JD[KD]] = np.nan
TEMP = N

f = interpolate.interp1d(time,hum,kind='nearest',bounds_error=False,fill_value=np.nan)
N = f(TIME)
JD = np.where(np.isnan(N)==0)[0]
KD = np.where(np.diff(N[JD])==0)
N[JD[KD]] = np.nan
HUM = N

f = interpolate.interp1d(time,pwi,kind='nearest',bounds_error=False,fill_value=np.nan)
N = f(TIME)
JD = np.where(np.isnan(N)==0)[0]
KD = np.where(np.diff(N[JD])==0)
N[JD[KD]] = np.nan
PWI = N

f = interpolate.interp1d(time,wis,kind='nearest',bounds_error=False,fill_value=np.nan)
N = f(TIME)
JD = np.where(np.isnan(N)==0)[0]
KD = np.where(np.diff(N[JD])==0)
N[JD[KD]] = np.nan
WIS = N

f = interpolate.interp1d(time,rain,kind='nearest',bounds_error=False,fill_value=np.nan)
N = f(TIME)
JD = np.where(np.isnan(N)==0)[0]
KD = np.where(np.diff(N[JD])==0)
N[JD[KD]] = np.nan
RAIN = N

f = interpolate.interp1d(time,pres,kind='nearest',bounds_error=False,fill_value=np.nan)
N = f(TIME)
JD = np.where(np.isnan(N)==0)[0]
KD = np.where(np.diff(N[JD])==0)
N[JD[KD]] = np.nan
PRES = N



################################################################ SAVE ########
# Locations: LOCS = ['BB','BO','HA','HT','LB','LBP','LBS']
# TIME: decimal years (daily increments, at midday)
# NUT1: NO3+NO2 (mg/L)
# NUT2: O-Phos (mg/L)
# NUT3: TN (mg/L)
# NUT4: T-Phos (mg/L)
# TOX1: LCMSMS Cylindro (ppb)
# TOX2: LCMSMS Microcystin (ppb)
# TOX3: ELISA Cylindro (ppb)
# TOX4: ELISA Microcystin (ppb)
# DIV: bacterial/algal family name
# DEN: algal concentration
# TBV: total biovolume
# FBV: fractional biovolume
# TEMP: temperature
# HUM: humidity
# PWI: peak wind speed
# WIS: wind speed
# RAIN: rain
# PRES: barometric pressure
np.savez("../Preprocessed/Data_historical_7day.npz",LOCS=LOCS,TIME=TIME,NUT1=NUT1,
        NUT2=NUT2,NUT3=NUT3,NUT4=NUT4,TOX1=TOX1,TOX2=TOX2,TOX3=TOX3,TOX4=TOX4,DIV=DIV,DEN=DEN,
        TBV=TBV,FBV=FBV,TEMP=TEMP,HUM=HUM,PWI=PWI,WIS=WIS,RAIN=RAIN,PRES=PRES)


