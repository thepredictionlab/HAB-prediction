## Code to preprocess Detroit Lake empirical data Spring 2019
#! Mathew Titus and James Watson, The Prediction Lab 2019

# exec(open('../Data/Scripts/make_latest.py').read())

import pandas as pd
import numpy as np
import datetime
import openpyxl as px
import sys
import os
import matplotlib.pylab as plt
from scipy import interpolate
interp_kind = 'zero' # 'previous'

###############################################################################
#### Function for calculating decimal years
from datetime import datetime
def yearLength(yr):
    yr = np.reshape(yr,(-1,1))
    year_length = [datetime(year=yr[k]+1, month=1, day=1) - datetime(year=yr[k], month=1, day=1) for k in np.arange(len(yr))]
    year_length = [year_length[k].days for k in np.arange(len(yr))]
    return year_length

def datetime2year(dt):
    year_part = dt - datetime(year=dt.year, month=1, day=1)
    year_length = datetime(year=dt.year+1, month=1, day=1) - datetime(year=dt.year, month=1, day=1)
    return dt.year + year_part/year_length

def year2datetime(date):
    year = np.array(np.uint(np.floor(date)), ndmin=2)
    year_part = date - year
    Seconds = year_part*np.uint(yearLength(year))*24*60*60
    Date = [ dt.datetime(year=year[0][k] ,month=1,day=1) + dt.timedelta(seconds = Seconds[0][k]) for k in np.arange(year.size) ]
    return Date

fileName = "03_01_19.xlsx"

# check for top directory level - 'HAB-prediction' used on GitHub, 'GitHub' used on my laptop.
# check for other top directories (containing 'BMA', 'Data' folders, etc.) by adding other try statements.
cpath = os.path.abspath('.').split('/')
try:
    fatherLevel = cpath.index('HAB-prediction')
except(ValueError):
    try:
        fatherLevel = cpath.index('GitHub')
    except(ValueError):
        raise Exception('Top directory not found.')

fatherDir = '/'.join(cpath[:fatherLevel+1]) + '/'
dataDir = fatherDir+'Data/Raw_lake/Latest/'
saveDir = fatherDir+'Data/Preprocessed/Latest/'
#### Load excel spreadsheet from Salem
data = px.load_workbook(dataDir+fileName,data_only=True)
ws = data['Sheet1']
faultLog = ''

alreadyRan = False
f = open(dataDir+'pastRuns.txt','r')
for line in f:
	print(line)
	if line == fileName:
		alreadyRan = True
		print("It's in there")
		f.close()
		sys.exit('Already ran.')
		# This data has been processed and added to the master files

# f.close()
# if ~alreadyRan:
# 	f = open(dataDir+'pastRuns.txt','a')
# 	f.write('\n'+fileName)
# 	f.close()

##################### Extract values ###########################


################## Cyanotoxins ###########################
# Date of toxin sample
column = ws['A']
time = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	date = column[x].value
	try:
		if len(str(date)) > 0:
			time.append(datetime2year(date))
	except:
		errors = ''.join((errors,' ',str(x)))
if len(errors)>0:
	error = 'Column A, failed to append rows '+errors+'.\n'
	faultLog = ''.join((faultLog,error))
time = np.array(time)
time = time[[k!=None for k in time]]

# Site of toxin sample
column = ws['C']
locs = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	loc = column[x].value
	try:
		if len(str(loc)) > 0:
			locs.append(loc)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column C, failed to append rows '+errors+'.\n'
	faultLog = ''.join((faultLog,error))
locs = np.array(locs)
locs = locs[[k!=None for k in locs]]

# Microcystin toxin level (ppb)
column = ws['E']
mics = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	mic = column[x].value
	try:
		if len(str(mic)) > 0:
			mics.append(mic)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column E, failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
mics = np.array(mics)
mics = mics[[k!=None for k in mics]]

# Cylindrospermopsin toxin level (ppb)
column = ws['F']
cyls = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	cyl = column[x].value
	try:
		if len(str(cyl)) > 0:
			cyls.append(cyl)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column F, failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
cyls = np.array(cyls)
cyls = cyls[[k!=None for k in cyls]]

#########################
### COMPOSE DATA FILE ###
#########################

np.savez(saveDir+'data_cyano.npz',TIME=time,LOCS=locs,MICRO=mics,CYLIN=cyls)

##########################
### UPDATE MASTER FILE ###
##########################

try:
	hCyano = np.load(saveDir+'hist_cyano.npz')
	hTime = np.vstack(( np.reshape(hCyano['TIME'],(-1,1)), np.reshape(time,(-1,1)) ))
	hLocs = np.vstack(( np.reshape(hCyano['LOCS'],(-1,1)), np.reshape(locs,(-1,1)) ))
	hMicro = np.vstack(( np.reshape(hCyano['MICRO'],(-1,1)), np.reshape(mics,(-1,1)) ))
	hCylin = np.vstack(( np.reshape(hCyano['CYLIN'],(-1,1)), np.reshape(cyls,(-1,1)) ))
	np.savez(saveDir+'hist_cyano.npz',TIME=hTime,LOCS=hLocs,MICRO=hMicro,CYLIN=hCylin)
except(FileNotFoundError):
	hCyano = np.load(saveDir+'data_cyano.npz')
	np.savez(saveDir+'hist_cyano.npz',TIME=hCyano['TIME'],LOCS=hCyano['LOCS'],\
		MICRO=hCyano['MICRO'],CYLIN=hCyano['CYLIN'])

################## Nutrients ###########################
# Date of sample
columnd = ws['H']
columnt = ws['I']
time = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	try:
		date1 = columnd[x].value
		date2 = columnt[x].value
		# account for missing times
		date = datetime.combine(date1,date2)
		if len(str(date)) > 0:
			time.append(datetime2year(date))
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column H or I failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
time = np.array(time)
time = time[[k!=None for k in time]]

# Site of nutrient sample
column = ws['J']
locs = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	loc = column[x].value
	try:
		if len(str(loc)) > 0:
			locs.append(loc)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column J failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
locs = np.array(locs)
locs = locs[[k!=None for k in locs]]

# Ammonium level
column = ws['L']
amms = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	amm = column[x].value
	try:
		if len(str(amm)) > 0:
			amms.append(amm)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column L failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
amms = np.array(amms)
amms = amms[[k!=None for k in amms]]

# Nitrate/nitrite level
column = ws['M']
nits = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	nit = column[x].value
	try:
		if len(str(nit)) > 0:
			nits.append(nit)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column M failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
nits = np.array(nits)
nits = nits[[k!=None for k in nits]]

# Organic phosphorous level
column = ws['N']
phos = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	pho = column[x].value
	try:
		if len(str(pho)) > 0:
			phos.append(pho)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column N failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
phos = np.array(phos)
phos = phos[[k!=None for k in phos]]

# Total nitrogen level
column = ws['O']
tns = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	tn = column[x].value
	try:
		if len(str(tn)) > 0:
			tns.append(tn)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column O failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
tns = np.array(tns)
tns = tns[[k!=None for k in tns]]

# Total phosphorous level
column = ws['P']
tps = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	tp = column[x].value
	try:
		if len(str(tp)) > 0:
			tps.append(tp)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column P failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
tps = np.array(tps)
tps = tps[[k!=None for k in tps]]

#########################
### COMPOSE DATA FILE ###
#########################

np.savez(saveDir+'data_nuts.npz',TIME=time,LOCS=locs,NH3=amms,NO=nits,OP=phos,\
	TN=tns,TP=tps)

##########################
### UPDATE MASTER FILE ###
##########################

try:
	hNuts = np.load(saveDir+'hist_nuts.npz')
	hTime = np.vstack(( np.reshape(hNuts['TIME'],(-1,1)), np.reshape(time,(-1,1)) ))
	hLocs = np.vstack(( np.reshape(hNuts['LOCS'],(-1,1)), np.reshape(locs,(-1,1)) ))
	hNH3 = np.vstack(( np.reshape(hNuts['NH3'],(-1,1)), np.reshape(amms,(-1,1)) ))
	hNO = np.vstack(( np.reshape(hNuts['NO'],(-1,1)), np.reshape(nits,(-1,1)) ))
	hOP = np.vstack(( np.reshape(hNuts['OP'],(-1,1)), np.reshape(phos,(-1,1)) ))
	hTN = np.vstack(( np.reshape(hNuts['TN'],(-1,1)), np.reshape(tns,(-1,1)) ))
	hTP = np.vstack(( np.reshape(hNuts['TP'],(-1,1)), np.reshape(tps,(-1,1)) ))
	np.savez(saveDir+'hist_nuts.npz',TIME=hTime,LOCS=hLocs,NH3=hNH3,NO=hNO,OP=hOP,\
		TN=hTN,TP=hTP)
except(FileNotFoundError):
	hNuts = np.load(saveDir+'data_nuts.npz')
	np.savez(saveDir+'hist_nuts.npz',TIME=hNuts['TIME'],LOCS=hNuts['LOCS'],\
		NH3=hNuts['NH3'],NO=hNuts['NO'],OP=hNuts['OP'],TN=hNuts['TN'],TP=hNuts['TP'])

################## YSI Data ###########################
# Date of sample
columnd = ws['R']
columnt = ws['S']
time = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	try:
		date1 = columnd[x].value
		date2 = columnt[x].value
		date = datetime.combine(date1,date2)
		if len(str(date)) > 0:
			time.append(datetime2year(date))
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column R or S failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
time = np.array(time)
time = time[[k!=None for k in time]]

# Site of YSI data sample
column = ws['T']
locs = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	loc = column[x].value
	try:
		if len(str(loc)) > 0:
			locs.append(loc)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column T failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
locs = np.array(locs)
locs = locs[[k!=None for k in locs]]

# Depth of sample
column = ws['U']
deps = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	dep = column[x].value
	try:
		if len(str(dep)) > 0:
			deps.append(dep)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column U failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
deps = np.array(deps)
deps = deps[[k!=None for k in deps]]

# Water temperature
column = ws['V']
tems = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	tem = column[x].value
	try:
		if len(str(tem)) > 0:
			tems.append(tem)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column V failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
tems = np.array(tems)
tems = tems[[k!=None for k in tems]]

# Barometer
column = ws['W']
bars = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	bar = column[x].value
	try:
		if len(str(bar)) > 0:
			bars.append(bar)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column W failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
bars = np.array(bars)
bars = bars[[k!=None for k in bars]]

# Percentage dissolved oxygen ---- w.r.t. what unit?
column = ws['X']
oxps = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	oxp = column[x].value
	try:
		if len(str(oxp)) > 0:
			oxps.append(oxp)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column X failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
oxps = np.array(oxps)
oxps = oxps[[k!=None for k in oxps]]

# Concentration of dissolved oxygen (mg/L)
column = ws['Y']
oxcs = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	oxc = column[x].value
	try:
		if len(str(oxc)) > 0:
			oxcs.append(oxc)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column Y failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
oxcs = np.array(oxcs)
oxcs = oxcs[[k!=None for k in oxcs]]

# Dissolved oxygen %L ---- what is this?
column = ws['Z']
dols = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	dol = column[x].value
	try:
		if len(str(dol)) > 0:
			dols.append(dol)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column Z failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
dols = np.array(dols)
dols = dols[[k!=None for k in dols]]

# SPC-uS/cm ---- what is this?
column = ws['AA']
spcs = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	spc = column[x].value
	try:
		if len(str(spc)) > 0:
			spcs.append(spc)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AA failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
spcs = np.array(spcs)
spcs = spcs[[k!=None for k in spcs]]

# pH
column = ws['AB']
phs = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	ph = column[x].value
	try:
		if len(str(loc)) > 0:
			phs.append(ph)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AB failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
phs = np.array(phs)
phs = phs[[k!=None for k in phs]]

# NTU ---- what is this?
column = ws['AC']
ntus = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	ntu = column[x].value
	try:
		if len(str(ntu)) > 0:
			ntus.append(ntu)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AC failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
ntus = np.array(ntus)
ntus = ntus[[k!=None for k in ntus]]

# BGA-PC RFU ---- what is this?
column = ws['AD']
bgas = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	bga = column[x].value
	try:
		if len(str(bga)) > 0:
			bgas.append(bga)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AD failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
bgas = np.array(bgas)
bgas = bgas[[k!=None for k in bgas]]

# BGA-PC ug/L
column = ws['AE']
bgals = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	bgal = column[x].value
	try:
		if len(str(bgal)) > 0:
			bgals.append(bgal)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AE failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
bgals = np.array(bgals)
bgals = bgals[[k!=None for k in bgals]]

# Chl RFU ---- what is this?
column = ws['AF']
chlus = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	chlu = column[x].value
	try:
		if len(str(chlu)) > 0:
			chlus.append(chlu)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AF failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
chlus = np.array(chlus)
chlus = chlus[[k!=None for k in chlus]]

# Chl ug/L
column = ws['AG']
chlls = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	chll = column[x].value
	try:
		if len(str(chll)) > 0:
			chlls.append(chll)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AG failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
chlls = np.array(chlls)
chlls = chlls[[k!=None for k in chlls]]

#########################
### COMPOSE DATA FILE ###
#########################

np.savez(saveDir+'data_ysi.npz',TIME=time,LOCS=locs,DEPTH=deps,TEMP=tems,PRES=bars,\
	DO_PCT=oxps,DO_CON=oxcs,DO_PCT_L=dols,SPC=spcs,PH=phs,NTU=ntus,BGA_RFU=bgas,\
	BGA_PC=bgals,CHL_RFU=chlus,CHL_CON=chlls)

##########################
### UPDATE MASTER FILE ###
##########################

try:
	hYSI = np.load(saveDir+'hist_ysi.npz')
	hTime = np.vstack(( np.reshape(hYSI['TIME'],(-1,1)), np.reshape(time,(-1,1)) ))
	hLocs = np.vstack(( np.reshape(hYSI['LOCS'],(-1,1)), np.reshape(locs,(-1,1)) ))
	hDeps = np.vstack(( np.reshape(hYSI['DEPTH'],(-1,1)), np.reshape(deps,(-1,1)) ))
	hTems = np.vstack(( np.reshape(hYSI['TEMP'],(-1,1)), np.reshape(tems,(-1,1)) ))
	hBars = np.vstack(( np.reshape(hYSI['PRES'],(-1,1)), np.reshape(bars,(-1,1)) ))
	hOxps = np.vstack(( np.reshape(hYSI['DO_PCT'],(-1,1)), np.reshape(oxps,(-1,1)) ))
	hOxcs = np.vstack(( np.reshape(hYSI['DO_CON'],(-1,1)), np.reshape(oxcs,(-1,1)) ))
	hDols = np.vstack(( np.reshape(hYSI['DO_PCT_L'],(-1,1)), np.reshape(dols,(-1,1)) ))
	hSpcs = np.vstack(( np.reshape(hYSI['SPC'],(-1,1)), np.reshape(spcs,(-1,1)) ))
	hPhs = np.vstack(( np.reshape(hYSI['PH'],(-1,1)), np.reshape(phs,(-1,1)) ))
	hNtus = np.vstack(( np.reshape(hYSI['NTU'],(-1,1)), np.reshape(ntus,(-1,1)) ))
	hBgas = np.vstack(( np.reshape(hYSI['BGA_RFU'],(-1,1)), np.reshape(bgas,(-1,1)) ))
	hBgals = np.vstack(( np.reshape(hYSI['BGA_PC'],(-1,1)), np.reshape(bgals,(-1,1)) ))
	hChlus = np.vstack(( np.reshape(hYSI['CHL_RFU'],(-1,1)), np.reshape(chlus,(-1,1)) ))
	hChlls = np.vstack(( np.reshape(hYSI['CHL_CON'],(-1,1)), np.reshape(chlls,(-1,1)) ))
	np.savez(saveDir+'hist_ysi.npz',TIME=hTime,LOCS=hLocs,DEPTH=hDeps,TEMP=hTems,PRES=hBars,\
		DO_PCT=hOxps,DO_CON=hOxcs,DO_PCT_L=hDols,SPC=hSpcs,PH=hPhs,NTU=hNtus,BGA_RFU=hBgas,\
		BGA_PC=hBgals,CHL_RFU=hChlus,CHL_CON=hChlls)
except(FileNotFoundError):
	hYSI = np.load(saveDir+'data_ysi.npz')
	np.savez(saveDir+'hist_ysi.npz',TIME=hYSI['TIME'],LOCS=hYSI['LOCS'],DEPTH=hYSI['DEPTH'],\
		TEMP=hYSI['TEMP'],PRES=hYSI['PRES'],DO_PCT=hYSI['DO_PCT'],DO_CON=hYSI['DO_CON'],\
		DO_PCT_L=hYSI['DO_PCT_L'],SPC=hYSI['SPC'],PH=hYSI['PH'],NTU=hYSI['NTU'],\
		BGA_RFU=hYSI['BGA_RFU'],BGA_PC=hYSI['BGA_PC'],CHL_RFU=hYSI['CHL_RFU'],CHL_CON=hYSI['CHL_CON'])

################## Algae Speciation ###########################
# Date of sample
columnd = ws['AI']
columnt = ws['AJ']
time = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	try:
		date1 = columnd[x].value
		date2 = columnt[x].value
		date = datetime.combine(date1,date2)
		if len(str(date)) > 0:
			time.append(datetime2year(date))
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AI or AJ failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
time = np.array(time)
time = time[[k!=None for k in time]]

# Site of algae data sample
column = ws['AK']
locs = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	loc = column[x].value
	try:
		if len(str(loc)) > 0:
			locs.append(loc)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AK failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
locs = np.array(locs)
locs = locs[[k!=None for k in locs]]

# Depth of algae sample
column = ws['AL']
deps = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	dep = column[x].value
	try:
		if len(str(dep)) > 0:
			deps.append(dep)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AL failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
deps = np.array(deps)
deps = deps[[k!=None for k in deps]]

# Sample aliquot (mL)
column = ws['AM']
alis = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	ali = column[x].value
	try:
		if len(str(ali)) > 0:
			alis.append(ali)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AM failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
alis = np.array(alis)
alis = alis[[k!=None for k in alis]]

# Field area ---- what is this?
column = ws['AN']
fars = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	far = column[x].value
	try:
		if len(str(far)) > 0:
			fars.append(far)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AN failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
fars = np.array(fars)
fars = fars[[k!=None for k in fars]]

# Fields ---- what is this?
column = ws['AO']
flds = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	fld = column[x].value
	try:
		if len(str(fld)) > 0:
			flds.append(fld)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AO failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
flds = np.array(flds)
flds = flds[[k!=None for k in flds]]

# Area counted ---- what is this?
column = ws['AP']
arcs = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	arc = column[x].value
	try:
		if len(str(arc)) > 0:
			arcs.append(arc)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AP failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
arcs = np.array(arcs)
arcs = arcs[[k!=None for k in arcs]]

# Factor1 ---- what is this?
column = ws['AQ']
facs = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	fac = column[x].value
	try:
		if len(str(fac)) > 0:
			facs.append(fac)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AQ failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
facs = np.array(facs)
facs = facs[[k!=None for k in facs]]

# Genus
column = ws['AR']
gens = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	gen = column[x].value
	try:
		if len(str(gen)) > 0:
			gens.append(gen)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AR failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
gens = np.array(gens)
gens = gens[[k!=None for k in gens]]

# Division
column = ws['AT']
divs = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	div = column[x].value
	try:
		if len(str(div)) > 0:
			divs.append(div)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AT failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
divs = np.array(divs)
divs = divs[[k!=None for k in divs]]

# Tally rep 1 ---- what is this?
column = ws['AU']
tlys = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	tly = column[x].value
	try:
		if len(str(tly)) > 0:
			tlys.append(tly)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AU failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
tlys = np.array(tlys)
tlys = tlys[[k!=None for k in tlys]]

# Density (cells/mL)
column = ws['AW']
dens = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	den = column[x].value
	try:
		if len(str(den)) > 0:
			dens.append(den)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AW failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
dens = np.array(dens)
dens = dens[[k!=None for k in dens]]

# Total biovolume (um3/mL)
column = ws['AY']
tbvs = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	tbv = column[x].value
	try:
		if len(str(tbv)) > 0:
			tbvs.append(tbv)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AY failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
tbvs = np.array(tbvs)
tbvs = tbvs[[k!=None for k in tbvs]]

# Percentage biovolume
column = ws['AZ']
fbvs = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	fbv = column[x].value
	try:
		if len(str(fbv)) > 0:
			fbvs.append(fbv)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column AZ failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
fbvs = np.array(fbvs)
fbvs = fbvs[[k!=None for k in fbvs]]

# Percentage abundance
column = ws['BA']
abus = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	abu = column[x].value
	try:
		if len(str(abu)) > 0:
			abus.append(abu)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column BA failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
abus = np.array(abus)
abus = abus[[k!=None for k in abus]]

#########################
### COMPOSE DATA FILE ###
#########################

np.savez(saveDir+'data_algae.npz',TIME=time,LOCS=locs,DEPTH=deps,ALIQUOT=alis,FIELD_AREA=fars,\
	FIELDS=flds,AREA=arcs,FACTOR=facs,GENUS=gens,DIV=divs,TALLY=tlys,DENSITY=dens,TBV=tbvs,\
	FBV=fbvs,ABUNDANCE=abus)

##########################
### UPDATE MASTER FILE ###
##########################

try:
	hAlgae = np.load(saveDir+'hist_algae.npz')
	hTime = np.vstack(( np.reshape(hAlgae['TIME'],(-1,1)), np.reshape(time,(-1,1)) ))
	hLocs = np.vstack(( np.reshape(hAlgae['LOCS'],(-1,1)), np.reshape(locs,(-1,1)) ))
	hDeps = np.vstack(( np.reshape(hAlgae['DEPTH'],(-1,1)), np.reshape(deps,(-1,1)) ))
	hAlis = np.vstack(( np.reshape(hAlgae['ALIQUOT'],(-1,1)), np.reshape(alis,(-1,1)) ))
	hFars = np.vstack(( np.reshape(hAlgae['FIELD_AREA'],(-1,1)), np.reshape(fars,(-1,1)) ))
	hFlds = np.vstack(( np.reshape(hAlgae['FIELDS'],(-1,1)), np.reshape(flds,(-1,1)) ))
	hArcs = np.vstack(( np.reshape(hAlgae['AREA'],(-1,1)), np.reshape(arcs,(-1,1)) ))
	hFacs = np.vstack(( np.reshape(hAlgae['FACTOR'],(-1,1)), np.reshape(facs,(-1,1)) ))
	hGens = np.vstack(( np.reshape(hAlgae['GENUS'],(-1,1)), np.reshape(gens,(-1,1)) ))
	hDivs = np.vstack(( np.reshape(hAlgae['DIV'],(-1,1)), np.reshape(divs,(-1,1)) ))
	hTlys = np.vstack(( np.reshape(hAlgae['TALLY'],(-1,1)), np.reshape(tlys,(-1,1)) ))
	hDens = np.vstack(( np.reshape(hAlgae['DENSITY'],(-1,1)), np.reshape(dens,(-1,1)) ))
	hTbvs = np.vstack(( np.reshape(hAlgae['TBV'],(-1,1)), np.reshape(tbvs,(-1,1)) ))
	hFbvs = np.vstack(( np.reshape(hAlgae['FBV'],(-1,1)), np.reshape(fbvs,(-1,1)) ))
	hAbus = np.vstack(( np.reshape(hAlgae['ABUNDANCE'],(-1,1)), np.reshape(abus,(-1,1)) ))
	np.savez(saveDir+'hist_algae.npz',TIME=hTime,LOCS=hLocs,DEPTH=hDeps,ALIQUOT=hAlis,FIELD_AREA=hFars,\
		FIELDS=hFlds,AREA=hArcs,FACTOR=hFacs,GENUS=hGens,DIV=hDivs,TALLY=hTlys,DENSITY=hDens,\
		TBV=hTbvs,FBV=hFbvs,ABUNDANCE=hAbus)
except(FileNotFoundError):
	hAlgae = np.load(saveDir+'data_algae.npz')
	np.savez(saveDir+'hist_algae.npz',TIME=hAlgae['TIME'],LOCS=hAlgae['LOCS'],DEPTH=hAlgae['DEPTH'],\
		ALIQUOT=hAlgae['ALIQUOT'],FIELD_AREA=hAlgae['FIELD_AREA'],FIELDS=hAlgae['FIELDS'],\
		AREA=hAlgae['AREA'],FACTOR=hAlgae['FACTOR'],GENUS=hAlgae['GENUS'],DIV=hAlgae['DIV'],\
		TALLY=hAlgae['TALLY'],DENSITY=hAlgae['DENSITY'],TBV=hAlgae['TBV'],FBV=hAlgae['FBV'],\
		ABUNDANCE=hAlgae['ABUNDANCE'])

################## Weather data @ Detroit Ranger Station ###########################
# Date of sample
column = ws['CB']
time = []
errors = ''
for x in np.arange(2,len(column)):
	date = column[x].value
	try:
		if len(str(date)) > 0:
			time.append(datetime2year(date))
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column CB failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
time = np.array(time)
time = time[[k!=None for k in time]]

# Temperature
column = ws['CC']
tems = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	tem = column[x].value
	try:
		if len(str(tem)) > 0:
			tems.append(tem)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column CC failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
tems = np.array(tems)
tems = tems[[k!=None for k in tems]]

# Relative humidity
column = ws['CD']
hums = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	hum = column[x].value
	try:
		if len(str(hum)) > 0:
			hums.append(hum)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column CD failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
hums = np.array(hums)
hums = hums[[k!=None for k in hums]]

# Dew point
column = ws['CE']
dews = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	dew = column[x].value
	try:
		if len(str(dew)) > 0:
			dews.append(dew)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column CE failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
dews = np.array(dews)
dews = dews[[k!=None for k in dews]]

# Wind direction
column = ws['CF']
wdirs = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	wdir = column[x].value
	try:
		if len(str(wdir)) > 0:
			wdirs.append(wdir)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column CF failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
wdirs = np.array(wdirs)
wdirs = wdirs[[k!=None for k in wdirs]]

# Peak wind gust
column = ws['CG']
pkwgs = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	pkwg = column[x].value
	try:
		if len(str(pkwg)) > 0:
			pkwgs.append(pkwg)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column CG failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
pkwgs = np.array(pkwgs)
pkwgs = pkwgs[[k!=None for k in pkwgs]]

# Wind speed
column = ws['CH']
wspds = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	wspd = column[x].value
	try:
		if len(str(wspd)) > 0:
			wspds.append(wspd)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column CH failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
wspds = np.array(wspds)
wspds = wspds[[k!=None for k in wspds]]

# Solar radiation (cumulative)
column = ws['CI']
sols = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	sol = column[x].value
	try:
		if len(str(sol)) > 0:
			sols.append(sol)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column CI failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
sols = np.array(sols)
sols = sols[[k!=None for k in sols]]

# Barometric pressure
column = ws['CJ']
barps = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	barp = column[x].value
	try:
		if len(str(barp)) > 0:
			barps.append(barp)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column CJ failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
barps = np.array(barps)
barps = barps[[k!=None for k in barps]]

# Precipitation
column = ws['CK']
precs = [] # np.zeros(len(column)-2)
errors = ''
for x in np.arange(2,len(column)):
	prec = column[x].value
	try:
		if len(str(prec)) > 0:
			precs.append(prec)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column CK failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
precs = np.array(precs)
precs = precs[[k!=None for k in precs]]

#########################
### COMPOSE DATA FILE ###
#########################

np.savez(saveDir+'data_weather.npz',TIME=time,TEMP=tems,HUM=hums,DEW=dews,WDIR=wdirs,PEAKW=pkwgs,\
	WSPD=wspds,SOLAR=sols,PRES=barps,PRECIP=precs)

##########################
### UPDATE MASTER FILE ###
##########################

try:
	hWeather = np.load(saveDir+'hist_weather.npz')
	hTime = np.vstack(( np.reshape(hWeather['TIME'],(-1,1)), np.reshape(time,(-1,1)) ))
	hTems = np.vstack(( np.reshape(hWeather['TEMP'],(-1,1)), np.reshape(tems,(-1,1)) ))
	hHums = np.vstack(( np.reshape(hWeather['HUM'],(-1,1)), np.reshape(hums,(-1,1)) ))
	hDews = np.vstack(( np.reshape(hWeather['DEW'],(-1,1)), np.reshape(dews,(-1,1)) ))
	hWdirs = np.vstack(( np.reshape(hWeather['WDIR'],(-1,1)), np.reshape(wdirs,(-1,1)) ))
	hPkwgs = np.vstack(( np.reshape(hWeather['PEAKW'],(-1,1)), np.reshape(pkwgs,(-1,1)) ))
	hWspds = np.vstack(( np.reshape(hWeather['WSPD'],(-1,1)), np.reshape(wspds,(-1,1)) ))
	hSols = np.vstack(( np.reshape(hWeather['SOLAR'],(-1,1)), np.reshape(sols,(-1,1)) ))
	hBarps = np.vstack(( np.reshape(hWeather['PRES'],(-1,1)), np.reshape(barps,(-1,1)) ))
	hPrecs = np.vstack(( np.reshape(hWeather['PRECIP'],(-1,1)), np.reshape(precs,(-1,1)) ))
	np.savez(saveDir+'hist_weather.npz',TIME=hTime,TEMP=hTems,HUM=hHums,DEW=hDews,WDIR=hWdirs,\
		PEAKW=hPkwgs,WSPD=hWspds,SOLAR=hSols,PRES=hBarps,PRECIP=hPrecs)
except(FileNotFoundError):
	hWeather = np.load(saveDir+'data_weather.npz')
	np.savez(saveDir+'hist_weather.npz',TIME=hWeather['TIME'],TEMP=hWeather['TEMP'],\
		HUM=hWeather['HUM'],DEW=hWeather['DEW'],WDIR=hWeather['WDIR'],PEAKW=hWeather['PEAKW'],\
		WSPD=hWeather['WSPD'],SOLAR=hWeather['SOLAR'],PRES=hWeather['PRES'],PRECIP=hWeather['PRECIP'])

################## USGS Data ###########################

###########################################################################
################## I ASSUME THE TIMES FOR EACH  ###########################
################## USGS DATA SET ARE IDENTICAL! ###########################
###########################################################################
# Date of sample
column = ws['CU']
time = []
errors = ''
for x in np.arange(2,len(column)):
	date = column[x].value
	try:
		if len(str(date)) > 0:
			time.append(datetime2year(date))
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column CU failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
time = np.array(time)
time = time[[k!=None for k in time]]

# Reservoir level - Big Cliff
column = ws['CV']
bclvls = []
errors = ''
for x in np.arange(2,len(column)):
	bclvl = column[x].value
	try:
		if len(str(bclvl)) > 0:
			bclvls.append(bclvl)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column CV failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
bclvls = np.array(bclvls)
bclvls = bclvls[[k!=None for k in bclvls]]


# # Date of sample
# column = ws['CW']
# time = []
# errors = ''
# for x in np.arange(2,len(column)):
# 	date = column[x].value
# 	try:
# 		if len(str(date)) > 0:
# 			time.append(datetime2year(date))
# 	except:
# 		errors = ''.join((errors,' ',str(x)))

# if len(errors)>0:
# 	error = 'Column CW failed to append rows '+str(errors)+'.\n'
# 	faultLog = ''.join((faultLog,error))
# precs = precs[[k!=None for k in precs]]

# Tailwater level - Big Cliff
column = ws['CX']
bctwls = []
errors = ''
for x in np.arange(2,len(column)):
	bctwl = column[x].value
	try:
		if len(str(bctwl)) > 0:
			bctwls.append(bctwl)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column CX failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
bctwls = np.array(bctwls)
bctwls = bctwls[[k!=None for k in bctwls]]

# # Date of sample
# column = ws['CY']
# time = []
# errors = ''
# for x in np.arange(2,len(column)):
# 	date = column[x].value
# 	try:
# 		if len(str(date)) > 0:
# 			time.append(datetime2year(date))
# 	except:
# 		errors = ''.join((errors,' ',str(x)))

# if len(errors)>0:
# 	error = 'Column CY failed to append rows '+str(errors)+'.\n'
# 	faultLog = ''.join((faultLog,error))
# precs = precs[[k!=None for k in precs]]

# Blowout Creek Discharge
column = ws['CZ']
bodis = []
errors = ''
for x in np.arange(2,len(column)):
	bodi = column[x].value
	try:
		if len(str(bodi)) > 0:
			bodis.append(bodi)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column CZ failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
bodis = np.array(bodis)
bodis = bodis[[k!=None for k in bodis]]

# Blowout Creek Level
column = ws['DA']
bolvls = []
errors = ''
for x in np.arange(2,len(column)):
	bolvl = column[x].value
	try:
		if len(str(bolvl)) > 0:
			bolvls.append(bolvl)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column DA failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
bolvls = np.array(bolvls)
bolvls = bolvls[[k!=None for k in bolvls]]

# Blowout Creek Water Temp
column = ws['DB']
bowts = []
errors = ''
for x in np.arange(2,len(column)):
	bowt = column[x].value
	try:
		if len(str(bowt)) > 0:
			bowts.append(bowt)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column DB failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
bowts = np.array(bowts)
bowts = bowts[[k!=None for k in bowts]]

# # Date of sample
# column = ws['DC']
# time = []
# errors = ''
# for x in np.arange(2,len(column)):
# 	date = column[x].value
# 	try:
# 		if len(str(date)) > 0:
# 			time.append(datetime2year(date))
# 	except:
# 		errors = ''.join((errors,' ',str(x)))

# if len(errors)>0:
# 	error = 'Column DC failed to append rows '+str(errors)+'.\n'
# 	faultLog = ''.join((faultLog,error))
# precs = precs[[k!=None for k in precs]]

# Breitenbush River Discharge
column = ws['DD']
bbdis = []
errors = ''
for x in np.arange(2,len(column)):
	bbdi = column[x].value
	try:
		if len(str(bbdi)) > 0:
			bbdis.append(bbdi)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column DD failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
bbdis = np.array(bbdis)
bbdis = bbdis[[k!=None for k in bbdis]]

# Breitenbush River Level
column = ws['DE']
bblvls = []
errors = ''
for x in np.arange(2,len(column)):
	bblvl = column[x].value
	try:
		if len(str(bblvl)) > 0:
			bblvls.append(bblvl)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column DE failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
bblvls = np.array(bblvls)
bblvls = bblvls[[k!=None for k in bblvls]]

# Breitenbush River Water Temp
column = ws['DF']
bbwts = []
errors = ''
for x in np.arange(2,len(column)):
	bbwt = column[x].value
	try:
		if len(str(bbwt)) > 0:
			bbwts.append(bbwt)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column DF failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
bbwts = np.array(bbwts)
bbwts = bbwts[[k!=None for k in bbwts]]

# # Date of sample
# column = ws['DG']
# time = []
# errors = ''
# for x in np.arange(2,len(column)):
# 	date = column[x].value
# 	try:
# 		if len(str(date)) > 0:
# 			time.append(datetime2year(date))
# 	except:
# 		errors = ''.join((errors,' ',str(x)))

# if len(errors)>0:
# 	error = 'Column DG failed to append rows '+str(errors)+'.\n'
# 	faultLog = ''.join((faultLog,error))
# precs = precs[[k!=None for k in precs]]

# Tailwater level - Detroit Dam
column = ws['DH']
ddtwls = []
errors = ''
for x in np.arange(2,len(column)):
	bclvl = column[x].value
	try:
		if len(str(bclvl)) > 0:
			bclvls.append(bclvl)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column DH failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
ddtwls = np.array(ddtwls)
ddtwls = ddtwls[[k!=None for k in ddtwls]]


# # Date of sample
# column = ws['DI']
# time = []
# errors = ''
# for x in np.arange(2,len(column)):
# 	date = column[x].value
# 	try:
# 		if len(str(date)) > 0:
# 			time.append(datetime2year(date))
# 	except:
# 		errors = ''.join((errors,' ',str(x)))

# if len(errors)>0:
# 	error = 'Column DI failed to append rows '+str(errors)+'.\n'
# 	faultLog = ''.join((faultLog,error))
# precs = precs[[k!=None for k in precs]]

# Reservoir level - Detroit Dam
column = ws['DJ']
ddlvls = []
errors = ''
for x in np.arange(2,len(column)):
	ddlvl = column[x].value
	try:
		if len(str(ddlvl)) > 0:
			ddlvls.append(ddlvl)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column DJ failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
ddlvls = np.array(ddlvls)
ddlvls = ddlvls[[k!=None for k in ddlvls]]

# # Date of sample
# column = ws['DK']
# time = []
# errors = ''
# for x in np.arange(2,len(column)):
# 	date = column[x].value
# 	try:
# 		if len(str(date)) > 0:
# 			time.append(datetime2year(date))
# 	except:
# 		errors = ''.join((errors,' ',str(x)))

# if len(errors)>0:
# 	error = 'Column DK failed to append rows '+str(errors)+'.\n'
# 	faultLog = ''.join((faultLog,error))
# precs = precs[[k!=None for k in precs]]

# Little North Santiam - Discharge
column = ws['DL']
lnsds = []
errors = ''
for x in np.arange(2,len(column)):
	lnsd = column[x].value
	try:
		if len(str(lnsd)) > 0:
			lnsds.append(lnsd)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column DL failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
lnsds = np.array(lnsds)
lnsds = lnsds[[k!=None for k in lnsds]]

# Little North Santiam - Level
column = ws['DM']
lnsls = []
errors = ''
for x in np.arange(2,len(column)):
	lnsl = column[x].value
	try:
		if len(str(lnsl)) > 0:
			lnsls.append(lnsl)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column DM failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
lnsls = np.array(lnsls)
lnsls = lnsls[[k!=None for k in lnsls]]

# # Date of sample
# column = ws['DN']
# time = []
# errors = ''
# for x in np.arange(2,len(column)):
# 	date = column[x].value
# 	try:
# 		if len(str(date)) > 0:
# 			time.append(datetime2year(date))
# 	except:
# 		errors = ''.join((errors,' ',str(x)))

# if len(errors)>0:
# 	error = 'Column DN failed to append rows '+str(errors)+'.\n'
# 	faultLog = ''.join((faultLog,error))
# precs = precs[[k!=None for k in precs]]

# North Santiam - Discharge
column = ws['DO']
nsds = []
errors = ''
for x in np.arange(2,len(column)):
	nsd = column[x].value
	try:
		if len(str(nsd)) > 0:
			nsds.append(nsd)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column DO failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
nsds = np.array(nsds)
nsds = nsds[[k!=None for k in nsds]]

# North Santiam - Level
column = ws['DP']
nsls = []
errors = ''
for x in np.arange(2,len(column)):
	nsl = column[x].value
	try:
		if len(str(nsl)) > 0:
			nsls.append(nsl)
	except:
		errors = ''.join((errors,' ',str(x)))

if len(errors)>0:
	error = 'Column DP failed to append rows '+str(errors)+'.\n'
	faultLog = ''.join((faultLog,error))
nsls = np.array(nsls)
nsls = nsls[[k!=None for k in nsls]]

########################################
### COMPLETE COLLECTION OF USGS DATA ###
########################################
# 2 More N Santiam Locs, and Willamette w/in Salem.

#########################
### COMPOSE DATA FILE ###
#########################

np.savez(saveDir+'data_usgs.npz',TIME=time,BC_RES=bclvls,BC_TAIL=bctwls,BO_DIS=bodis,\
	BO_LVL=bolvls,BO_TEM=bowts,BB_DIS=bbdis,BB_LVL=bblvls,BB_TEM=bbwts,DD_TAIL=ddtwls,\
	DD_LVL=ddlvls,LNS_DIS=lnsds,LNS_LVL=lnsls,NS_DIS=nsds,NS_LVL=nsls)

##########################
### UPDATE MASTER FILE ###
##########################

try:
	hUsgs = np.load(saveDir+'hist_usgs.npz')
	hTime = np.vstack(( np.reshape(hUsgs['TIME'],(-1,1)), np.reshape(time,(-1,1)) ))
	hBclvls = np.vstack(( np.reshape(hUsgs['BC_RES'],(-1,1)), np.reshape(bclvls,(-1,1)) ))
	hBctwls = np.vstack(( np.reshape(hUsgs['BC_TAIL'],(-1,1)), np.reshape(bctwls,(-1,1)) ))
	hBodis = np.vstack(( np.reshape(hUsgs['BO_DIS'],(-1,1)), np.reshape(bodis,(-1,1)) ))
	hBolvls = np.vstack(( np.reshape(hUsgs['BO_LVL'],(-1,1)), np.reshape(bolvls,(-1,1)) ))
	hBowts = np.vstack(( np.reshape(hUsgs['BO_TEM'],(-1,1)), np.reshape(bowts,(-1,1)) ))
	hBbdis = np.vstack(( np.reshape(hUsgs['BB_DIS'],(-1,1)), np.reshape(bbdis,(-1,1)) ))
	hBblvls = np.vstack(( np.reshape(hUsgs['BB_LVL'],(-1,1)), np.reshape(bblvls,(-1,1)) ))
	hBbwts = np.vstack(( np.reshape(hUsgs['BB_TEM'],(-1,1)), np.reshape(bbwts,(-1,1)) ))
	hDdtwls = np.vstack(( np.reshape(hUsgs['DD_TAIL'],(-1,1)), np.reshape(ddtwls,(-1,1)) ))
	hDdlvls = np.vstack(( np.reshape(hUsgs['DD_LVL'],(-1,1)), np.reshape(ddlvls,(-1,1)) ))
	hLnsds = np.vstack(( np.reshape(hUsgs['LNS_DIS'],(-1,1)), np.reshape(lnsds,(-1,1)) ))
	hLnsls = np.vstack(( np.reshape(hUsgs['LNS_LVL'],(-1,1)), np.reshape(lnsls,(-1,1)) ))
	hNsds = np.vstack(( np.reshape(hUsgs['NS_DIS'],(-1,1)), np.reshape(nsds,(-1,1)) ))
	hNsls = np.vstack(( np.reshape(hUsgs['NS_LVL'],(-1,1)), np.reshape(nsls,(-1,1)) ))
	np.savez(saveDir+'hist_usgs.npz',TIME=hTime,BC_RES=hBclvls,BC_TAIL=hBctwls,BO_DIS=hBodis,\
		BO_LVL=hBolvls,BO_TEM=hBowts,BB_DIS=hBbdis,BB_LVL=hBblvls,BB_TEM=hBbwts,DD_TAIL=hDdtwls,\
		DD_LVL=hDdlvls,LNS_DIS=hLnsds,LNS_LVL=hLnsls,NS_DIS=hNsds,NS_LVL=hNsls)
except(FileNotFoundError):
	hUsgs = np.load(saveDir+'data_usgs.npz')
	np.savez(saveDir+'hist_usgs.npz',TIME=hUsgs['TIME'],BC_RES=hUsgs['BC_RES'],\
		BC_TAIL=hUsgs['BC_TAIL'],BO_DIS=hUsgs['BO_DIS'],BO_LVL=hUsgs['BO_LVL'],\
		BO_TEM=hUsgs['BO_TEM'],BB_DIS=hUsgs['BB_DIS'],BB_LVL=hUsgs['BB_LVL'],\
		BB_TEM=hUsgs['BB_TEM'],DD_TAIL=hUsgs['DD_TAIL'],DD_LVL=hUsgs['DD_LVL'],\
		LNS_DIS=hUsgs['LNS_DIS'],LNS_LVL=hUsgs['LNS_LVL'],NS_DIS=hUsgs['NS_DIS'],\
		NS_LVL=hUsgs['NS_LVL'])

################################################################################


f = open(dataDir+'pastRuns.txt','a')
f.write(fileName+'\n')
f.close()

f = open(fatherDir+'Data/Raw_lake/Latest/log.txt','w')
f.write(faultLog)
f.close()






